from DrissionPage import Chromium
from openpyxl import Workbook

wb=Workbook()
ws=wb.active

web = Chromium().latest_tab

web.get('https://hotels.ctrip.com/hotels/detail/?cityEnName=Nanning&cityId=380&hotelId=12645878&checkIn=2025-09-19&checkOut=2025-09-20&adult=1&children=0&crn=1&ages=&curr=CNY&barcurr=CNY&masterhotelid_tracelogid=100053755-0a6262dc-488406-99507&detailFilters=17%7C1~17~1*80%7C2~80~2*29%7C1~29~1%7C1&hotelType=normal&display=incavg&subStamp=630&isCT=true&isFirstEnterDetail=T')
# web.wait.eles_loaded('/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div/div[2]/div[1]/div[3]/div')
web.wait(1)
web('x:/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div/div[2]/div[1]/div[3]/div').click()
web.listen.start(targets='getHotelCommentList',is_regex=False,method='POST')

current_row=1
for page in range(1,6):
    web.wait.eles_loaded('x://*[@id="hp_container"]/div[2]/div[3]/div/div[2]/div[1]/div[4]/div[1]/div[2]')
    packet=web.listen.wait(timeout=30)
    comm=packet.response.body
    # for commentL in comm['data']['commentList']:
    #     print(f'用户ID:{commentL["id"]},用户评论:{commentL["content"]},评论图片:{commentL["imageList"]},用户评分:{commentL["rating"]},客服回复:{commentL["feedbackList"][0]['content']}')

    if page==1:
        ws.cell(row=current_row,column=1,value="用户ID")
        ws.cell(row=current_row,column=2,value="用户评分")
        ws.cell(row=current_row,column=3,value="用户评论")
        ws.cell(row=current_row,column=4,value="评论图片")
        ws.cell(row=current_row,column=5,value="客服回复")
        current_row+=1
    for commentL in comm['data']['commentList']:
        #可能出现不存在的情况
        feedback=commentL['feedbackList'][0]['content'] if commentL['feedbackList'] else ''
        if 'imageList' not in commentL:
            image=''
        else:
            image=str(commentL['imageList'])
        #写入excel表格
        ws.cell(row=current_row,column=1,value=commentL['id'])
        ws.cell(row=current_row,column=2,value=commentL['rating'])
        ws.cell(row=current_row,column=3,value=commentL['content'])
        ws.cell(row=current_row,column=4,value=image)
        ws.cell(row=current_row,column=5,value=feedback)
        current_row+=1
    web.wait.eles_loaded('x:/html/body/div[2]/div[2]/div[3]/div/div[2]/div[1]/div[4]/ul/li[9]/a/i')
    web('x:/html/body/div[2]/div[2]/div[3]/div/div[2]/div[1]/div[4]/ul/li[9]/a/i').click()

ws.title='携程酒店评论'
wb.save('携程酒店评论.xlsx')
print('携程酒店评论.xlsx保存成功')